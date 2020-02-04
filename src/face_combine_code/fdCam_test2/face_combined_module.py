import cv2
import os
import numpy as np
from PIL import Image
import sqlite3
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import RPi.GPIO as GPIO

def servo_motor1():
    try:
        pin1 = 5
        GPIO.setwarnings(False)
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(pin1, GPIO.OUT)

        p=GPIO.PWM(pin1,50)
        p.start(0)
        print(" 1 motor drive")
        print(" 180 degree")
        p.ChangeDutyCycle(12)
        time.sleep(2)

        print(" 0 degree")
        p.ChangeDutyCycle(2.5)
        time.sleep(2)
    except KeyboardInterrupt:
        p.stop()
        GPIO.cleanup()

def servo_motor2():
    try:
        pin2 = 6
        GPIO.setwarnings(False)
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(pin2, GPIO.OUT)

        p=GPIO.PWM(pin2,50)
        p.start(0)
        print(" 2 motor drive")
        print(" 180 degree")
        p.ChangeDutyCycle(12)
        time.sleep(2)

        print(" 0 degree")
        p.ChangeDutyCycle(2.5)
        time.sleep(2)
    except KeyboardInterrupt:
        p.stop()
        GPIO.cleanup()

def servo_motor3():
    try:
        pin3 = 19
        GPIO.setwarnings(False)
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(pin3, GPIO.OUT)

        p=GPIO.PWM(pin3,50)
        p.start(0)
        print(" 3 motor drive")
        print(" 180 degree")
        p.ChangeDutyCycle(12)
        time.sleep(2)

        print(" 0 degree")
        p.ChangeDutyCycle(2.5)
        time.sleep(2)
    except KeyboardInterrupt:
        p.stop()
        GPIO.cleanup()

def create_database():
    # conn = sqlite3.connect('./db/patient.db')
    # cur = conn.cursor()
    # cur.execute('drop table if exists list')
    # cur.execute('create table list( id int, name char(20))')
    # conn.commit()
    # cur.close()
    # conn.close()
    write_wb = Workbook()
    write_wb.save('./db/patient.xlsx')

#def append_xl():

def patient_information():
    wb = openpyxl.load_workbook('./db/patient.xlsx')
    sheet1 = wb.active
    face_id = int(input("\n(1). id를 입력하세요. ex_)0,1,2.. <return> ==>  "))
    name = input("\n(2). 이름을 입력하세요. ==>")
    medicine1 = int(input("\n(3). 1번약 투약 횟수 ==>"))
    medicine2 = int(input("\n(4). 2번약 투약 횟수 ==>"))
    medicine3 = int(input("\n(5). 3번약 투약 횟수 ==>"))

    sheet1.append([face_id, name, medicine1, medicine2, medicine3])
    wb.save('./db/patient.xlsx')
    wb.close()


def collecting_data():
    #============ sqlite3 db에 저장==============#

    # conn = sqlite3.connect('./db/patient.db')
    # cur = conn.cursor()   
    # face_id = input("\n id를 입력하세요. ex_)0,1,2.. <return> ==>  ")
    # name = input("\n 이름을 입력하세요. ==>")
    # cur.execute('insert into list(id, name) values(?,?)',(int(face_id), name))
    # conn.commit()
    # cur.execute("select * from list")
    # rows = cur.fetchall()
    # print(" id          name  ")
    # print("====================")
    # for row in rows:
    #     print(" {}          {}".format(row[0],row[1]))
    # cur.close()
    # conn.close()

    #============ 엑셀에 저장 =================#

     
    wb = openpyxl.load_workbook('./db/patient.xlsx', data_only = True)
    sheet1 = wb.active

    id_number = int(input("학습시킬 환자의 id를 입력하세요 ex_)0,1,2.. <return> ==>"))
    id_number = id_number+1
    face_id=sheet1.cell( row=id_number, column=1).value
    print("환자 번호 = ", face_id)

    
    cam = cv2.VideoCapture(0)
    cam.set(3, 640) # set video width
    cam.set(4, 480) # set video height
    face_detector = cv2.CascadeClassifier('haarcascades/haarcascade_frontalface_default.xml')
    print("\n [INFO] Initializing face capture, Look the camera and wait ...")

    count = 0
    while(True):
        ret, img = cam.read()
        img = cv2.flip(img, -1)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_detector.detectMultiScale(gray, 1.3, 5)
        for(x,y,w,h) in faces:
            cv2.rectangle(img, (x,y), (x+w,y+h), (255,0,0), 2)
            count+=1
            cv2.imwrite("dataset/User." +str(face_id) +'.' +str(count) + ".jpg", gray[y : y+h, x: x+w]) # user.1.0.jpg
        cv2.imshow('image', img)
        
        k = cv2.waitKey(100) & 0xff
        if k == 27:
            break
        elif count >=50:
            break
    print("\n [INFO] Exiting Program and cleanup stuff")
    cam.release()
    cv2.destroyAllWindows()


def training_data():
    path = 'dataset'
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    detector = cv2.CascadeClassifier("haarcascades/haarcascade_frontalface_default.xml")
    def getImagesAndLabels(path):
        imagePaths = [os.path.join(path,f) for f in os.listdir(path)] # imagepath를 리스트로 저장
        faceSamples = []
        ids = []
        for imagePath in imagePaths: # ex_) dataset//0.jpg
            PIL_img = Image.open(imagePath).convert('L') # convert it to grayscale
            img_numpy = np.array(PIL_img, 'uint8') # array 배열로 전환
            id = int(os.path.split(imagePath)[-1].split(".")[1])
            faces = detector.detectMultiScale(img_numpy)
            for (x,y,w,h) in faces:
                faceSamples.append(img_numpy[y : y+h, x: x+w])
                ids.append(id) #
        return faceSamples, ids
    print("\n [INFO] Training faces. It will take a few seconds. Wait ...")
    faces, ids = getImagesAndLabels(path) #리스트로 faces, ids에 들어감 
    recognizer.train(faces, np.array(ids)) # faces = faceSamples[], ids = ids[]
    recognizer.write('trainer/trainer.yml') 
    print("\n [INFO] {0} faces trained. Exiting Program".format(len(np.unique(ids))))


def recognizing_data():
    # conn = sqlite3.connect('./db/patient.db')
    # cur = conn.cursor()   
    # cur.execute("select * from list")
    # rows = cur.fetchall()
    # names=[]
    # for row in rows:
    #     names.append(row[1])
    # print("등록된 환자 =",names)
    # cur.close()
    # conn.close()

    names=[]
    wb = openpyxl.load_workbook('./db/patient.xlsx', data_only = True)
    sheet1 = wb.active

    for i in sheet1.rows:
        names.append(i[1].value)

    print(names)
    

    person1 = True
    person2 = True

    recognizer = cv2.face.LBPHFaceRecognizer_create() # recognizer (train, write, read, predict)
    recognizer.read('trainer/trainer.yml')
    cascadePath="haarcascades/haarcascade_frontalface_default.xml"
    faceCascade = cv2.CascadeClassifier(cascadePath) # face detector
    font = cv2.FONT_HERSHEY_SIMPLEX
    #id = 0
    cam = cv2.VideoCapture(0)
    cam.set(3, 640) # set video width
    cam.set(4, 480) # set video height
    minW = 0.1*cam.get(3)
    minH = 0.1*cam.get(4)
    count_et = 0
    while True:
        #count_et = count_et+1
        #print(count_et)
        ret, img = cam.read() # img를 계속 읽어와라
        img = cv2.flip(img, -1) # Flip vertically
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        faces = faceCascade.detectMultiScale(
            gray,
            scaleFactor = 1.2,
            minNeighbors = 5,
            minSize = (int(minW), int(minH)),
        )           
        for(x,y,w,h) in faces:
            cv2.rectangle(img, (x,y), (x+w, y+h), (0,255,0), 2)
            id, confidence = recognizer.predict(gray[y: y+h, x: x+w]) # confidence : 틀릴 확률 # predict에 따라서 id에 맞게 들어감
            # Check if confidence is less then 100 ==> "0" is perfect match
            if (confidence <100): # 맞을 확률이 있으면
                if person1 == True:
                    if id == 0:
                        print("jane")
                        medicine1 = sheet1.cell( row=1, column=3).value
                        medicine2 = sheet1.cell( row=1, column=4).value
                        medicine3 = sheet1.cell( row=1, column=5).value

                        for i in range (medicine1):
                            servo_motor1()
                        for i in range (medicine2):
                            servo_motor2()
                        for i in range (medicine3):
                            servo_motor3()
                        
                        person1 = False
                        #run_stop()
                if person2 == True:
                    if id ==1:
                        print("susan")
                        medicine1 = sheet1.cell( row=2, column=3).value
                        medicine2 = sheet1.cell( row=2, column=4).value
                        medicine3 = sheet1.cell( row=2, column=5).value

                        for i in range (medicine1):
                            servo_motor1()
                        for i in range (medicine2):
                            servo_motor2()
                        for i in range (medicine3):
                            servo_motor3()                        
                           
                        person2 = False
                        #run_stop()
                id = names[id]
                confidence2 = int(round(100-confidence))
                confidence = "  {0}%".format(round(100 - confidence))
                # if confidence2 >=30:
                #     for i in range(100):
                #         p.ChangeDutyCycle(i)
                #         time.sleep(0.5)                    
            else :
                id = "unknown"
                #confidence3 = int(round(confidence))
                confidence = "  {0}%".format(round(100 - confidence)) # 틀릴 확률이 100을 넘어가기 때문에 
                # choice = input("미등록된 사용자가 감지되었습니다. 등록하시겠습니까? [y/n]")
                # if choice == "y":
                #     cam.release()
                #     cv2.destroyAllWindows()
                #     collecting_data()
                #     break
                # else:
                #     continue
                
            cv2.putText(img, str(id), (x+5,y-5), font, 1, (255,255,255), 2)
            cv2.putText(img, str(confidence), (x+5,y+h-5), font, 1, (255,255,0), 1)

        #cv2.imshow('camera', img)
        k = cv2.waitKey(10) & 0xff # Press 'ESC' for exiting video
        if k == 27:
            break
    wb.close()
    print("\n [INFO] Exiting Program and cleanup stuff")
    

def run_stop():
        try:
            pin = 21
            GPIO.setwarnings(False)
            GPIO.setmode(GPIO.BCM)
            GPIO.setup(pin, GPIO.OUT)

            GPIO.output(pin, True)
            time.sleep(1)
            GPIO.output(pin, False)
            time.sleep(1)

        except KeyboardInterrupt:
            GPIO.cleanup()

